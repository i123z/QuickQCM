from pdf2image import convert_from_path
import os
import cv2
import imutils
import easyocr
import json
import numpy as np
import openpyxl
import sys



relative_path = os.getcwd()
# Récupérer le chemin absolu du fichier en cours d'exécution
chemin_fichier = os.path.abspath(__file__)

# Récupérer le chemin absolu du dossier contenant le fichier
chemin_dossier = os.path.dirname(chemin_fichier)


nom_pdf = sys.argv[1]
nom_json = sys.argv[2]


print("relative_path" +  chemin_dossier)
with open(chemin_dossier + '\\' + nom_json, 'r') as f:
    # Charger les données JSON
    data_json = json.load(f)

print(data_json)




def getKey1(tuples):
    return tuples[1]

def getKey0(tuples):
    return tuples[0]

def case_cauchee(case):
      taille = len(case)*len(case[0])
      n = int((sum( [ sum(case[i]) for i in range(len(case))] ) /(taille * 255))*100)
      if n > 50:
            return True
      return False

def choix_etudiant(L):

    Liste = ["A", "B", "C", "D"]
    Choix = []
    for i in range(4):
        if L[i]:
                Choix.append(Liste[i])
    return Choix



def charger_reponses(path):
    global chemin_dossier
    """
    Convert the images from the path to a list of images. Save the images in the local directory.
    @param path - the path to the images
    @return the number of pages
    """
    images = convert_from_path(path)
    nombreDePages = len(images)
    for i in range(nombreDePages):
            images[i].save(f"{chemin_dossier}/copies/pages{str(i)}.jpg", 'JPEG')
    return nombreDePages

def nombre_pages_par_etudiant(nbrQuestion, nbrPages):
    pages_etudiant = 1
    if nbrQuestion <= 6:
            return pages_etudiant
    pages_etudiant = pages_etudiant + int(np.ceil((nbrQuestion-6)/7))
    if (nbrPages%pages_etudiant) != 0:
        return "Problem - Nombre de pages"
    return pages_etudiant


NOMBRE_PAGES = charger_reponses(chemin_dossier + "\\" + nom_pdf)
NIVEAU = data_json['niveau']
BAREM = data_json['notation']
NOMBRE_QUESTIONS = len(data_json['questions'])
NOMBRE_PAGES_ETUDIANT = nombre_pages_par_etudiant(NOMBRE_QUESTIONS, NOMBRE_PAGES)
REPONSES_CORRECTES = []
NOMBRE_ETUDIANTS = int(NOMBRE_PAGES/NOMBRE_PAGES_ETUDIANT)
PROFESSEUR = data_json['prof']
MATIERE = data_json['matiere']

CHOIX = ['A', 'B', 'C', 'D']
for Question in data_json["questions"]:
     REPONSES_CORRECTES.append( CHOIX[int(Question['reponse']) -1 ])

print(NOMBRE_PAGES, NIVEAU, BAREM, NOMBRE_QUESTIONS, NOMBRE_PAGES_ETUDIANT , NOMBRE_ETUDIANTS)

print(REPONSES_CORRECTES)

    
    



def extraire_reponses_page(img, identifiant):
    Reponses_etudiant = []
    if identifiant:
        bloc_id = img[250:460,1062:]

        bloc_id = cv2.GaussianBlur(bloc_id, (5, 5), 0)
        bloc_id =  cv2.threshold(bloc_id, 0, 255,cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
        reader = easyocr.Reader(['en'])
        bounds = reader.readtext(bloc_id, allowlist = '0123456789')
        id=bounds[1][1]
        Reponses_etudiant.append(id)
        print(id)
        img = img[460:,:]
 
    Rep = []
    #blurred = cv2.GaussianBlur(img, (5, 5), 0)
    edged = cv2.Canny(img, 75, 200)

    thresh = cv2.threshold(img, 0, 255,cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1] + edged


    # trouver les contours dans l'image seuillée, puis initialiser
    # la liste des contours qui correspondent aux questions
    # trouver les contours dans l'image seuillée, puis initialiser
    # la liste des contours qui correspondent aux questions
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)


    cnts = imutils.grab_contours(cnts)

    questionCnts = []
    LA=[] 
    # boucle sur les contours
    for c in cnts:
    # calcule la boite englobante du contour, puis utilise la
        # boîte englobante pour dériver le rapport d'aspect
        (x, y, w, h) = cv2.boundingRect(c)
        ar = w / float(h)

        if ar <4.5 and ar > 4:
            LA.append((x,y,w,h))
            questionCnts.append(c) 




    LA=sorted(LA, key = getKey1)
    #if identifiant:
    #    LA.pop(0)
    
    i=0
    for R in LA:
        (x,y,w,h) = R
        threshb=thresh[y:,x+140:][int(h/2):h,:w-150]


        thresha=thresh[y:,x+140:][:int(h/2),:w-150]


        cntsa = cv2.findContours(thresha.copy(), cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
        cntsa = imutils.grab_contours(cntsa)
        cntsb = cv2.findContours(threshb.copy(), cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
        cntsb = imutils.grab_contours(cntsb) 
        questiona = []
        questionb = []
        for c in cntsa:
            # calcule la boite englobante du contour, puis utilise la
                # boîte englobante pour dériver le rapport d'aspect
                (x, y, w, h) = cv2.boundingRect(c)
                ar = w / float(h)
                if  w >= 8 and h >= 8 and ar >= 0.5 and ar <= 2:
                    #:#
                    #LA.append((x, y, w, h))
                    #res = cv2.drawContours(image, [c], -1, (255, 0, 255), 2)
                    #cv2.imwrite("contours_or.png", image)
                    questiona.append((x,y,w,h))

                    
        questiona =sorted(questiona, key = getKey0)
        ChoixA = [e for e in questiona[1::2]]
        
        ChoixA_Bool = []
        for R in ChoixA:
                    (x,y,w,h) = R
                    region = thresha[y + 4 :y+h -4, x + 4:x+w - 4]

                    ChoixA_Bool.append(case_cauchee(region))
        ChoixA_alphabets = choix_etudiant(ChoixA_Bool)


            

                    
        for c in cntsb:
            # calcule la boite englobante du contour, puis utilise la
                # boîte englobante pour dériver le rapport d'aspect
                (x, y, w, h) = cv2.boundingRect(c)
                ar = w / float(h)
                if  w >= 8 and h >= 8 and ar >= 0.5 and ar <= 2:
                    #:#
                    #LA.append((x, y, w, h))
                    #res = cv2.drawContours(image, [c], -1, (255, 0, 255), 2)
                    #cv2.imwrite("contours_or.png", image)
                    questionb.append((x,y,w,h))

                    
        questionb =sorted(questionb, key = getKey0)
        ChoixB = [e for e in questionb[1::2]]

        ChoixB_Bool = []
        for R in ChoixB:
                    (x,y,w,h) = R
                    region = threshb[y + 4 :y+h - 4, x + 4:x+w - 4]

                    ChoixB_Bool.append(case_cauchee(region))
        ChoixB_alphabets = choix_etudiant(ChoixB_Bool)
        
        Reponses_etudiant.append((ChoixA_alphabets , ChoixB_alphabets))
                    
    return Reponses_etudiant



def extraire_reponses_etudiant(Nombre_de_pages, Numero_de_copie):
    Reponses = []
    Premiere_page_identifiant = True
    for i in range(Numero_de_copie*Nombre_de_pages, (Numero_de_copie + 1)*Nombre_de_pages):
        page = chemin_dossier + "\\copies/pages" + str(i) + ".jpg"
        img = cv2.imread(page,cv2.IMREAD_GRAYSCALE)
        Reponses += extraire_reponses_page(img, Premiere_page_identifiant)
        Premiere_page_identifiant = False
    reponses_dict = {}
    reponses_dict["Identifiant"] = Reponses[0]
    for i in range(1,len(Reponses)):
        reponses_dict["Question" + str(i)] = Reponses[i]
    return reponses_dict


def corriger_question(Reponses_question: tuple, Bonne_reponse: str, Barem: list):
    Reponse_correcte = Barem[0]
    Reponse_non_correcte = Barem[1]
    ChoixA = Reponses_question[0]
    ChoixB = Reponses_question[1]
    if len(ChoixB) == 0:
        if len(ChoixA) == 4 or len(ChoixA) == 0:
            return 0
        elif len(ChoixA) != 1:
            return Reponse_non_correcte
        else:
            if ChoixA[0] == Bonne_reponse:
                return Reponse_correcte
            else:
                return Reponse_non_correcte
    elif len(ChoixB) == 4:
        return 0
    elif len(ChoixB) != 1:
        return Reponse_non_correcte
    else:
        if ChoixB[0] == Bonne_reponse:
            return Reponse_correcte
        else:
            return Reponse_non_correcte



def Corriger_copie(Reponses_etudiant , Reponses_correctes, Barem):
    note = 0
    for i in range(len(Reponses_correctes)):
         Question = "Question" + str(i+1)
         note += corriger_question(Reponses_etudiant[Question], Reponses_correctes[i], Barem)
    return (Reponses_etudiant["Identifiant"] , note, Reponses_etudiant)


ID_NOTE_REPONSES = []

for i in range(NOMBRE_ETUDIANTS):
    reponses_etudiant = extraire_reponses_etudiant(NOMBRE_PAGES_ETUDIANT,i)
    ID_NOTE_REPONSES.append(Corriger_copie(reponses_etudiant, REPONSES_CORRECTES, BAREM))
    

print(ID_NOTE_REPONSES)


def Generer_fichier_excel_notes(id_note_reponses, niveau, matiere, nom_professeur ):
    # Ouvrir le fichier Excel
    dossier_fichiers_excel = chemin_dossier + '\\Listes etudiants excel/'
    nom_fichier_excel_notes = niveau + '.xlsx'
    nom_fichier_excel_emails = niveau + ' - Email.xlsx'
    wb_notes = openpyxl.load_workbook(dossier_fichiers_excel + nom_fichier_excel_notes)
    wb_emails = openpyxl.load_workbook(dossier_fichiers_excel + nom_fichier_excel_emails)


    # Accéder à la feuille de calcul
    feuille_notes = wb_notes['JM']
    feuille_emails = wb_emails['JM']
    print('nombre de lignes = ', feuille_emails.max_row)
    indice_debut = 7
    case_id = "E"
    case_note = 'D'
    nombre_etudiants = 20

    # Nnregistrement des notes dans le fichier excel
    feuille_notes['B4'] = matiere
    feuille_notes['C4'] = "Professeur: " + nom_professeur
    for i in range(nombre_etudiants):
        cellule_id = case_id + str(indice_debut + i)
        cellule_note = case_note + str(indice_debut + i)
        for (id, note, reponses) in id_note_reponses:
            if(str(feuille_emails[cellule_id].value) == id):
                feuille_notes[cellule_note] = note
                feuille_notes['G' + str(indice_debut + i)] = ""
    
    # Enregistrer les modifications dans un nouveau fichier
    fichier_excel_sortie = data_json["nomFichier"] + '.xlsx'
    wb_notes.save(dossier_fichiers_excel + fichier_excel_sortie)
    return fichier_excel_sortie



    
Generer_fichier_excel_notes(ID_NOTE_REPONSES, NIVEAU, MATIERE, PROFESSEUR )